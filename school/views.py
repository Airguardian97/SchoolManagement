from django.shortcuts import render,redirect,reverse
from . import forms,models,modelsv2
from django.db.models import Sum
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required,user_passes_test
from django.conf import settings
from django.core.mail import send_mail
# from school.modelsv2 import *
# from school.models import *
from django.db.models import Value, F, CharField
from django.db.models.functions import Concat
from django.http import JsonResponse 
from django.http import HttpResponse
from django.contrib.auth import logout  # Import the logout function
from django.db import IntegrityError
import re  # Add this import
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from django.contrib import messages
from exam import models as QMODEL
from students import models as SMODEL
from django.db.models import Max
from django.db.models import OuterRef, Subquery
from django.db.models import Prefetch
import socket
from Crypto.Cipher import AES  # Ensure AES is imported
from Crypto.Util.Padding import pad  # Ensure pad is imported from Crypto.Util.Padding
from base64 import b64encode


def home_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return render(request,'school/index.html')



#for showing signup/login button for teacher
def adminclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return render(request,'school/adminclick.html')


#for showing signup/login button for teacher
def teacherclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return render(request,'school/teacherclick.html')


#for showing signup/login button for student
def studentclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return render(request,'school/studentclick.html')





def admin_signup_view(request):
    form=forms.AdminSigupForm()
    if request.method=='POST':
        form=forms.AdminSigupForm(request.POST)
        if form.is_valid():
            user=form.save()
            user.set_password(user.password)
            user.save()


            my_admin_group = Group.objects.get_or_create(name='ADMIN')
            my_admin_group[0].user_set.add(user)

            return HttpResponseRedirect('adminlogin')
    return render(request,'school/adminsignup.html',{'form':form})




from django.contrib.auth.models import User, Group
from openpyxl import load_workbook
from django.contrib import messages
from django.shortcuts import render, redirect
from . import forms

def bulk_signup_view(request):
    # Initialize the forms for rendering
    form1 = forms.StudentUserForm()
    form2 = forms.StudentExtraForm()    
    
    
    

    if request.method == 'POST':
        print("POST request received")

        # Check if the file is in FILES
        if request.FILES.get('csv_file'):
            print("File received")

            excel_file = request.FILES['csv_file']
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                print(f"Invalid file type received: {excel_file.name}")
                messages.error(request, "Please upload a valid Excel file.")
                return redirect('admin-add-student')  # redirect to the same page

            try:
                # Load the Excel file
                print(f"Loading Excel file: {excel_file.name}")
                wb = load_workbook(excel_file)
                sheet = wb.active

                # Iterate over the rows starting from the second row (to skip header)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    print("Processing row:", row)

                    if None in row or len(row) < 9:
                        print(f"Invalid data in row: {row}")
                        raise ValueError(f"Invalid data in row: {row}")

                    first_name, last_name, username, email, password, roll, mobile, student_class, fee = row

                    # Check for existing username or email
                    if User.objects.filter(username=username).exists():
                        print(f"Username '{username}' already exists.Skipping.")
                        continue  
                    if User.objects.filter(email=email).exists():
                        print(f"Email '{email}' already exists.Skipping.")
                        continue  

                    # Use StudentUserForm to create the user
                    form1 = forms.StudentUserForm({
                        'username': username,
                        'first_name': first_name,
                        'last_name': last_name,
                        'email': email,
                        'password': password
                    })

                    # Validate and save the user form
                    if form1.is_valid():
                        user = form1.save()
                        user.set_password(user.password)  # Hash the password
                        user.save()

                        # Use StudentExtraForm for additional student details
                        form2 = forms.StudentExtraForm({
                            'roll': roll,
                            'mobile': mobile,
                            'cl': student_class,
                            'fee': fee,
                            'status': "1"
                        })
                        # print(form2)
                        # Validate and save the extra student form
                        if form2.is_valid():
                            f2 = form2.save(commit=False)
                            f2.user = user  # Link the student extra data with the user
                            f2.save()

                            # Add the user to the student group
                            my_student_group, _ = Group.objects.get_or_create(name='STUDENT')
                            my_student_group.user_set.add(user)

                        else:
                            print(f"Form2 invalid for row: {row}")
                            raise ValueError("Invalid student data.")

                    else:
                        print(f"Form1 invalid for row: {row}")
                        raise ValueError("Invalid user data.")

                # Success message after processing all rows
                print("Bulk student registration successful!")
                messages.success(request, "Bulk student registration successful!")
                return redirect('afterlogin')

            except Exception as e:
                print(f"Error occurred: {str(e)}")
                messages.error(request, f"An error occurred: {str(e)}")
                return redirect('admin-add-student')

    return render(request, 'school/studentsignup.html', {'form1': form1, 'form2': form2})








def student_signup_view(request):
    form1=forms.StudentUserForm()
    form2=forms.StudentExtraForm()
    mydict={'form1':form1,'form2':form2}
    if request.method=='POST':
        form1=forms.StudentUserForm(request.POST)
        form2=forms.StudentExtraForm(request.POST)
        if form1.is_valid() and form2.is_valid():
            user=form1.save()
            user.set_password(user.password)
            user.save()
            f2=form2.save(commit=False)
            f2.user=user
            user2=f2.save()

            my_student_group = Group.objects.get_or_create(name='STUDENT')
            my_student_group[0].user_set.add(user)

        return HttpResponseRedirect('studentlogin')
    return render(request,'school/studentsignup.html',context=mydict)

# def bulk_upload_view(request):
#     if request.method == "POST":
#         # Handle form submission
#         form1 = StudentForm1(request.POST)
#         form2 = StudentForm2(request.POST)
#         if form1.is_valid() and form2.is_valid():
#             form1.save()
#             form2.save()
#             messages.success(request, "Student successfully registered!")
#             return redirect("bulk_upload")

#         # Handle bulk upload
#         if "excel_file" in request.FILES:
#             excel_file = request.FILES["excel_file"]

#             try:
#                 # Open the uploaded Excel file
#                 wb = load_workbook(excel_file)
#                 sheet = wb.active

#                 # Iterate through rows (adjust column mapping as needed)
#                 for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
#                     first_name, last_name, username, email, password, mobile, student_class, fee, roll = row

#                     # Create Student and Parent objects
#                     student = Student.objects.create(
#                         first_name=first_name,
#                         last_name=last_name,
#                         username=username,
#                         email=email,
#                         password=password,
#                         student_class=student_class,
#                         roll=roll,
#                         fee=fee
#                     )
#                     Parent.objects.create(
#                         student=student,
#                         mobile=mobile
#                     )

#                 messages.success(request, "Students uploaded successfully!")
#             except Exception as e:
#                 messages.error(request, f"An error occurred: {e}")
#                 return redirect("bulk_upload")
    
#     else:
#         form1 = StudentForm1()
#         form2 = StudentForm2()

#     return render(request, "school/admin_bulk_upload.html", {"form1": form1, "form2": form2})



def teacher_signup_view(request):
    form1=forms.TeacherUserForm()
    form2=forms.TeacherExtraForm()
    mydict={'form1':form1,'form2':form2}
    if request.method=='POST':
        form1=forms.TeacherUserForm(request.POST)
        form2=forms.TeacherExtraForm(request.POST)
        if form1.is_valid() and form2.is_valid():
            user=form1.save()
            user.set_password(user.password)
            user.save()
            f2=form2.save(commit=False)
            f2.user=user
            user2=f2.save()

            my_teacher_group = Group.objects.get_or_create(name='TEACHER')
            my_teacher_group[0].user_set.add(user)

        return HttpResponseRedirect('teacherlogin')
    return render(request,'school/teachersignup.html',context=mydict)






#for checking user is techer , student or admin
def is_admin(user):    
    return user.groups.filter(name='ADMIN').exists()
def is_teacher(user):
    return user.groups.filter(name='TEACHER').exists()
def is_student(user):
    return user.groups.filter(name='STUDENT').exists()
def is_student_email(user):
    return modelsv2.Parent.objects.filter(email_address=user.email).exists()
    
def afterlogin_view(request):
    try:
        if is_admin(request.user):
            return redirect('admin-dashboard')
        elif is_teacher(request.user):
            accountapproval = models.TeacherExtra.objects.all().filter(user_id=request.user.id, status=True)
            if accountapproval:
                return redirect('teacher-dashboard')
            else:
                return render(request, 'school/teacher_wait_for_approval.html')
        elif is_student(request.user):
            if is_student_email(request.user):
                accountapproval = models.StudentExtra.objects.all().filter(user_id=request.user.id, status=True)
                if accountapproval:
                    return redirect('student-dashboard')
                else:
                    return render(request, 'school/student_wait_for_approval.html')
            logout(request)          
            return HttpResponse("Error: Unauthorized access. Please contact support.", status=403)

    except Exception as e:
        # Log out the user if an error occurs
           # Log out the user if an error occurs
        logout(request)
        
        return HttpResponse(f"An error occurred: {str(e)}. You have been logged out for security reasons.", status=500)



#for dashboard of adminnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnn

@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    teachercount=models.TeacherExtra.objects.all().filter(status=True).count()
    pendingteachercount=models.TeacherExtra.objects.all().filter(status=False).count()

    studentcount=models.StudentExtra.objects.all().filter(status=True).count()
    pendingstudentcount=models.StudentExtra.objects.all().filter(status=False).count()

    teachersalary=models.TeacherExtra.objects.filter(status=True).aggregate(Sum('salary'))
    pendingteachersalary=models.TeacherExtra.objects.filter(status=False).aggregate(Sum('salary'))

    studentfee=models.StudentExtra.objects.filter(status=True).aggregate(Sum('fee',default=0))
    pendingstudentfee=models.StudentExtra.objects.filter(status=False).aggregate(Sum('fee'))

    notice=models.Notice.objects.all()

    #aggregate function return dictionary so fetch data from dictionay
    mydict={
        'teachercount':teachercount,
        'pendingteachercount':pendingteachercount,

        'studentcount':studentcount,
        'pendingstudentcount':pendingstudentcount,

        'teachersalary':teachersalary['salary__sum'],
        'pendingteachersalary':pendingteachersalary['salary__sum'],

        'studentfee':studentfee['fee__sum'],
        'pendingstudentfee':pendingstudentfee['fee__sum'],

        'notice':notice

    }

    return render(request,'school/admin_dashboard.html',context=mydict)







#for teacher sectionnnnnnnn by adminnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnn

@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_teacher_view(request):
    return render(request,'school/admin_teacher.html')

@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_add_teacher_view(request):
    form1=forms.TeacherUserForm()
    form2=forms.TeacherExtraForm()
    mydict={'form1':form1,'form2':form2}
    if request.method=='POST':
        form1=forms.TeacherUserForm(request.POST)
        form2=forms.TeacherExtraForm(request.POST)
        if form1.is_valid() and form2.is_valid():
            user=form1.save()
            user.set_password(user.password)
            user.save()

            f2=form2.save(commit=False)
            f2.user=user
            f2.status=True
            f2.save()

            my_teacher_group = Group.objects.get_or_create(name='TEACHER')
            my_teacher_group[0].user_set.add(user)

        return HttpResponseRedirect('admin-teacher')
    return render(request,'school/admin_add_teacher.html',context=mydict)


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_view_teacher_view(request):
    teachers=models.TeacherExtra.objects.all().filter(status=True)
    return render(request,'school/admin_view_teacher.html',{'teachers':teachers})


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_approve_teacher_view(request):
    teachers=models.TeacherExtra.objects.all().filter(status=False)
    return render(request,'school/admin_approve_teacher.html',{'teachers':teachers})


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def approve_teacher_view(request,pk):
    teacher=models.TeacherExtra.objects.get(id=pk)
    teacher.status=True
    teacher.save()
    return redirect(reverse('admin-approve-teacher'))


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def delete_teacher_view(request,pk):
    teacher=models.TeacherExtra.objects.get(id=pk)
    user=models.User.objects.get(id=teacher.user_id)
    user.delete()
    teacher.delete()
    return redirect('admin-approve-teacher')


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def delete_teacher_from_school_view(request,pk):
    teacher=models.TeacherExtra.objects.get(id=pk)
    user=models.User.objects.get(id=teacher.user_id)
    user.delete()
    teacher.delete()
    return redirect('admin-view-teacher')


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def update_teacher_view(request,pk):
    teacher=models.TeacherExtra.objects.get(id=pk)
    user=models.User.objects.get(id=teacher.user_id)

    form1=forms.TeacherUserForm(instance=user)
    form2=forms.TeacherExtraForm(instance=teacher)
    mydict={'form1':form1,'form2':form2}

    if request.method=='POST':
        form1=forms.TeacherUserForm(request.POST,instance=user)
        form2=forms.TeacherExtraForm(request.POST,instance=teacher)
        print(form1)
        if form1.is_valid() and form2.is_valid():
            user=form1.save()
            user.set_password(user.password)
            user.save()
            f2=form2.save(commit=False)
            f2.status=True
            f2.save()
            return redirect('admin-view-teacher')
    return render(request,'school/admin_update_teacher.html',context=mydict)


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_view_teacher_salary_view(request):
    teachers=models.TeacherExtra.objects.all()
    return render(request,'school/admin_view_teacher_salary.html',{'teachers':teachers})






#for student by adminnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnn

@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_student_view(request):
    return render(request,'school/admin_student.html')


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_add_student_view(request):
    form1=forms.StudentUserForm()
    form2=forms.StudentExtraForm()
    mydict={'form1':form1,'form2':form2}
    if request.method=='POST':
        form1=forms.StudentUserForm(request.POST)
        form2=forms.StudentExtraForm(request.POST)
        if form1.is_valid() and form2.is_valid():
            print("form is valid")
            user=form1.save()
            user.set_password(user.password)
            user.save()

            f2=form2.save(commit=False)
            f2.user=user
            f2.status=True
            f2.save()

            my_student_group = Group.objects.get_or_create(name='STUDENT')
            my_student_group[0].user_set.add(user)
        else:
            print("form is invalid")
        return HttpResponseRedirect('admin-student')
    return render(request,'school/admin_add_student.html',context=mydict)


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_view_student_view(request):
    students=models.StudentExtra.objects.all().filter(status=True)
    return render(request,'school/admin_view_student.html',{'students':students})


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def delete_student_from_school_view(request,pk):
    student=models.StudentExtra.objects.get(id=pk)
    user=models.User.objects.get(id=student.user_id)
    user.delete()
    student.delete()
    return redirect('admin-view-student')


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def delete_student_view(request,pk):
    student=models.StudentExtra.objects.get(id=pk)
    user=models.User.objects.get(id=student.user_id)
    user.delete()
    student.delete()
    return redirect('admin-approve-student')


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def update_student_view(request,pk):
    student=models.StudentExtra.objects.get(id=pk)
    user=models.User.objects.get(id=student.user_id)
    form1=forms.StudentUserForm(instance=user)
    form2=forms.StudentExtraForm(instance=student)
    mydict={'form1':form1,'form2':form2}
    if request.method=='POST':
        form1=forms.StudentUserForm(request.POST,instance=user)
        form2=forms.StudentExtraForm(request.POST,instance=student)
        print(form1)
        if form1.is_valid() and form2.is_valid():
            user=form1.save()
            user.set_password(user.password)
            user.save()
            f2=form2.save(commit=False)
            f2.status=True
            f2.save()
            return redirect('admin-view-student')
    return render(request,'school/admin_update_student.html',context=mydict)



@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_approve_student_view(request):
    students=models.StudentExtra.objects.all().filter(status=False)
    return render(request,'school/admin_approve_student.html',{'students':students})


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def approve_student_view(request,pk):
    students=models.StudentExtra.objects.get(id=pk)
    students.status=True
    students.save()
    return redirect(reverse('admin-approve-student'))


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_view_student_fee_view(request):
    students=models.StudentExtra.objects.all()
    return render(request,'school/admin_view_student_fee.html',{'students':students})






#attendance related viewwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwww
@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_attendance_view(request):
    return render(request,'school/admin_attendance.html')


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_take_attendance_view(request,cl):
    students=models.StudentExtra.objects.all().filter(cl=cl)
    print(students)
    aform=forms.AttendanceForm()
    if request.method=='POST':
        form=forms.AttendanceForm(request.POST)
        if form.is_valid():
            Attendances=request.POST.getlist('present_status')
            date=form.cleaned_data['date']
            for i in range(len(Attendances)):
                AttendanceModel=models.Attendance()
                AttendanceModel.cl=cl
                AttendanceModel.date=date
                AttendanceModel.present_status=Attendances[i]
                AttendanceModel.roll=students[i].roll
                AttendanceModel.save()
            return redirect('admin-attendance')
        else:
            print('form invalid')
    return render(request,'school/admin_take_attendance.html',{'students':students,'aform':aform})



@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_view_attendance_view(request,cl):
    form=forms.AskDateForm()
    if request.method=='POST':
        form=forms.AskDateForm(request.POST)
        if form.is_valid():
            date=form.cleaned_data['date']
            attendancedata=models.Attendance.objects.all().filter(date=date,cl=cl)
            studentdata=models.StudentExtra.objects.all().filter(cl=cl)
            mylist=zip(attendancedata,studentdata)
            return render(request,'school/admin_view_attendance_page.html',{'cl':cl,'mylist':mylist,'date':date})
        else:
            print('form invalid')
    return render(request,'school/admin_view_attendance_ask_date.html',{'cl':cl,'form':form})









#fee related view by adminnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnnn
@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_fee_view(request):
    return render(request,'school/admin_fee.html')


@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_view_fee_view(request,cl):
    feedetails=models.StudentExtra.objects.all().filter(cl=cl)
    return render(request,'school/admin_view_fee.html',{'feedetails':feedetails,'cl':cl})








#notice related viewsssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss
@login_required(login_url='adminlogin')
@user_passes_test(is_admin)
def admin_notice_view(request):
    form=forms.NoticeForm()
    if request.method=='POST':
        form=forms.NoticeForm(request.POST)
        if form.is_valid():
            form=form.save(commit=False)
            form.by=request.user.first_name
            form.save()
            return redirect('admin-dashboard')
    return render(request,'school/admin_notice.html',{'form':form})








#for TEACHER  LOGIN    SECTIONNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNNN
@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_dashboard_view(request):
    teacherdata=models.TeacherExtra.objects.all().filter(status=True,user_id=request.user.id)
    notice=models.Notice.objects.all()
    mydict={
        'salary':teacherdata[0].salary,
        'mobile':teacherdata[0].mobile,
        'date':teacherdata[0].joindate,
        'notice':notice
    }
    return render(request,'school/teacher_dashboard.html',context=mydict)



@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_attendance_view(request):
    teacher_id = str(request.user.id)
    teacher_id = "NT202401"
    subjects = modelsv2.Subject.objects.filter(teacher_id=teacher_id)
    print(subjects)
    return render(request, 'school/teacher_attendance.html', {
        'subjects': subjects
    })



@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_take_attendance_view(request, cl):
    # Fetch enrolled sr_id values from the Studentenrollsubject model
    enrolled_ids = modelsv2.Studentenrollsubject.objects.filter(subject_code=cl).values_list('sr_id', flat=True)

    # Fetch the Studentregister records where sr_id is in the enrolled_ids list
    students_register = modelsv2.Studentregister.objects.filter(sr_id__in=enrolled_ids)
    print(students_register)
    
    # Fetch the Student records where studentid is in the list of students from Studentregister
    students = modelsv2.Student.objects.filter(ref__in=students_register.values_list('stud_id', flat=True))
    print(students)

    # Ensure that there are students before proceeding
    if not students:
        # Handle case when there are no students in the class
        print("No students found for this class")
        return render(request, 'school/teacher_take_attendance.html', {
            'message': 'No students found for this class.',
            'cl': cl
        })

    aform = forms.AttendanceForm()

    if request.method == 'POST':
        form = forms.AttendanceForm(request.POST)
        if form.is_valid():
            Attendances = request.POST.getlist('present_status')
            date = form.cleaned_data['date']
            
            # Ensure that the number of attendance entries matches the number of students
            if len(Attendances) != len(students):
                print(f"Attendance count mismatch. Students: {len(students)}, Attendance entries: {len(Attendances)}")
                return render(request, 'school/teacher_take_attendance.html', {
                    'form': form,
                    'students': students,
                    'message': 'The number of attendance entries does not match the number of students.'
                })
            
            # Save attendance records
            for i in range(len(Attendances)):
                AttendanceModel = models.Attendance()
                AttendanceModel.cl = cl
                AttendanceModel.date = date
                AttendanceModel.present_status = Attendances[i]
                AttendanceModel.roll = students[i].ref  # or use another unique identifier if needed
                AttendanceModel.subject_code = cl  # Set the subject_code field
                AttendanceModel.save()
                
                        # After saving, send TCP message
                send_tcp_message_to_vb_server("0987654321", "Attendance saved successfully.")

            return redirect('teacher-attendance')
        else:
            print('Form invalid')
            # Log the form errors for debugging purposes
            print(form.errors)

    return render(request, 'school/teacher_take_attendance.html', {
        'students': students,
        'aform': aform,
        'cl': cl
    })


# AES Key and IV (must match the server!)
AES_KEY = b"1234567890123456"  # 16 bytes for AES-128
AES_IV = b"6543210987654321"  # 16 bytes for AES-128

def encrypt_message(message: str) -> str:
    """Encrypt the message using AES (CBC mode) and return the base64 encoded encrypted message."""
    
    # Initialize AES cipher in CBC mode with the specified key and IV
    cipher = AES.new(AES_KEY, AES.MODE_CBC, AES_IV)
    
    # Pad the message to ensure it's a multiple of the AES block size
    padded_data = pad(message.encode(), AES.block_size)
    
    # Encrypt the padded message
    encrypted_data = cipher.encrypt(padded_data)
    
    # Return the base64 encoded encrypted message
    return b64encode(encrypted_data).decode()


def send_tcp_message_to_vb_server(phone_number, message, host='172.16.1.15', port=5566):
    """Encrypt the message and send it to the VB.NET server over TCP."""
    try:
        encrypted_message = encrypt_message(f"{phone_number} -|- {message}")  # Encrypt the message
        
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.connect((host, port))  # Connect to the server
            s.sendall(encrypted_message.encode())  # Send the encrypted message (already base64 encoded)
            print("Sent via Server:", encrypted_message)
    
    except Exception as e:
        print("TCP Error:", e)
        
        
        
@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_view_attendance_view(request, cl):
    form = forms.AskDateForm()  # Initialize the form
    if request.method == 'POST':
        form = forms.AskDateForm(request.POST)
        if form.is_valid():
            # Get the date from the form
            date = form.cleaned_data['date']
            
            # Fetch the SR_ID values for the enrolled students in the given class (cl)
            enrolled_ids = modelsv2.Studentenrollsubject.objects.filter(subject_code=cl).values_list('sr_id', flat=True)
            
            # Fetch the Studentregister records for these enrolled students
            students_register = modelsv2.Studentregister.objects.filter(sr_id__in=enrolled_ids)
            
            # Fetch the Student records based on the studentid from Studentregister
            students = modelsv2.Student.objects.filter(ref__in=students_register.values_list('stud_id', flat=True))
            print(students)
            # Fetch the attendance data for the specified date and class
            attendancedata = models.Attendance.objects.filter(date=date, cl=cl)
            
            # Combine attendance data and student data using zip
            mylist = zip(attendancedata, students)

            # Render the template with the required context
            return render(request, 'school/teacher_view_attendance_page.html', {
                'cl': cl,
                'mylist': mylist,
                'date': date
            })
        else:
            print('Form is invalid')
    
    # Render the template for asking the date if the request is GET or the form is invalid
    return render(request, 'school/teacher_view_attendance_ask_date.html', {
        'cl': cl,
        'form': form
    })



@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_notice_view(request):
    form=forms.NoticeForm()
    if request.method=='POST':
        form=forms.NoticeForm(request.POST)
        if form.is_valid():
            form=form.save(commit=False)
            form.by=request.user.first_name
            form.save()
            return redirect('teacher-dashboard')
        else:
            print('form invalid')
    return render(request,'school/teacher_notice.html',{'form':form})







#FOR STUDENT AFTER THEIR Loginnnnnnnnnnnnnnnnnnnnn
@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_dashboard_view(request):
    studentdata=models.StudentExtra.objects.all().filter(status=True,user_id=request.user.id)
    notice=models.Notice.objects.all()
    mydict={
        'roll':studentdata[0].roll,
        'mobile':studentdata[0].mobile,
        'fee':studentdata[0].fee,
        'notice':notice
    }
    return render(request,'school/student_dashboard.html',context=mydict)



@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_attendance_view(request):
    form=forms.AskDateForm()
    if request.method=='POST':
        form=forms.AskDateForm(request.POST)
        if form.is_valid():
            date=form.cleaned_data['date']
            studentdata=models.StudentExtra.objects.all().filter(user_id=request.user.id,status=True)
            attendancedata=models.Attendance.objects.all().filter(date=date,cl=studentdata[0].cl,roll=studentdata[0].roll)
            mylist=zip(attendancedata,studentdata)
            return render(request,'school/student_view_attendance_page.html',{'mylist':mylist,'date':date})
        else:
            print('form invalid')
    return render(request,'school/student_view_attendance_ask_date.html',{'form':form})


import logging


@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_soa(request):
    user_email = request.user.email
    students = []
    selected_student = None
    student_info = {}
    transactions = []
    net_balance = 0.0

    try:
        # Fetch parent based on user email
        parents = modelsv2.Parent.objects.filter(email_address=user_email)

        if not parents.exists():
            print("Parent not found for the given email.")
            return render(request, 'school/error.html', {'message': 'Parent not found.'})

        parent = parents.first()  # Select the first parent if multiple exist for now

        # Fetch associated students
        parent_students = modelsv2.Parentstudent.objects.filter(gid=parent.pid)

        if parent_students.exists():
            students = modelsv2.Student.objects.filter(ref__in=[ps.stud_id for ps in parent_students])
            if students.exists():
                print("Students found:", students)
            else:
                print("No students found for this parent.")
        else:
            print("No student records associated with this parent.")

        # Get the selected student from the GET request
        student_ref = request.GET.get('student_ref')

        # If there's a selected student, filter by the selected student reference
        if student_ref:
            selected_student = students.get(ref=student_ref)  # This will raise an exception if not found
            print("Selected student:", selected_student)
        else:
            selected_student = students.first()  # Set the first student as default if none selected

        # Fetch the latest Studentregister data for the selected student
        latest_register = None
        if selected_student:
            latest_register = (
                SMODEL.Studentregister.objects.filter(stud_id=selected_student.ref)
                .annotate(latest_date=Max('dor'))  # Replace 'dor' with the correct field name
                .order_by('-latest_date')
                .first()
            )
            print("Latest register:", latest_register)

        # Prepare student information including grade level from Studentregister
        if selected_student:
            student_info = {
                'lrn_no': selected_student.lrn_no,
                'name': f"{selected_student.last_name}, {selected_student.first_name} {selected_student.middle_name or ''}",
                'grade': latest_register.grade_level if latest_register else "N/A",
                'voucher': selected_student.if_voucher,
            }

            # Retrieve charges and payments
            start_date = '2024-06-01'
            end_date = '2024-12-01'

            queryset_charges = modelsv2.Scharges.objects.filter(stud_id=selected_student.ref, date__range=[start_date, end_date])
            queryset_payments = modelsv2.Spayment.objects.filter(stud_id=selected_student.ref, date__range=[start_date, end_date])

            # Combine and prepare transactions for display
            charges = queryset_charges.annotate(
                c_amount=F('amount'),
                p_amount=Value(0),
                transaction_type=Value('Charge')
            )
            payments = queryset_payments.annotate(
                c_amount=Value(0),
                p_amount=F('amount'),
                transaction_type=Value('Payment')
            )

            transactions = sorted(
                list(charges) + list(payments),
                key=lambda t: t.date
            )

            # Calculate running balance
            balance = 0
            for transaction in transactions:
                balance += transaction.c_amount - transaction.p_amount
                transaction.balance = balance

            # Set the final net balance
            net_balance = balance

    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")
        return render(request, 'school/error.html', {'message': 'An unexpected error occurred.'})

    # Return the data to the template
    return render(request, 'school/student_soa.html', {
        'students': students,
        'selected_student': selected_student,
        'student_info': student_info,
        'transactions': transactions,
        'net_balance': net_balance,
    })

    
    
@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_list(request):
    user_email = request.user.email
    students = []
    student_info = []
    transactions = []
    net_balance = 0.0

    try:
        # Fetch parent based on user email
        parents = modelsv2.Parent.objects.filter(email_address=user_email)

        if not parents.exists():
            print("Parent not found for the given email.")
            return render(request, 'school/error.html', {'message': 'Parent not found.'})

        parent = parents.first()  # Select the first parent if multiple exist for now
        
        # Fetch associated students
        parent_students = modelsv2.Parentstudent.objects.filter(gid=parent.pid)
        print(parent_students)
        if parent_students.exists():
            students = modelsv2.Student.objects.filter(ref__in=[ps.stud_id for ps in parent_students])
            if students.exists():
                student_ids = students.values_list('ref', flat=True)

                print("Students found:", students)

                # Subquery for the latest Studentregister entry
                from django.db.models import OuterRef, Subquery

                latest_register_subquery = SMODEL.Studentregister.objects.filter(
                    stud_id=OuterRef('ref')
                ).order_by('-dor').values('pk')[:1]

                # Annotate the latest Studentregister entry
                students = students.annotate(latest_register_id=Subquery(latest_register_subquery))

                # Fetch additional Studentregister details
                student_info = [
                    {
                        'student': student,
                        'latest_register': SMODEL.Studentregister.objects.filter(pk=student.latest_register_id).first(),
                    }
                    for student in students
                ]

                # Process student_info to include grade level
                for info in student_info:
                    student_register = info['latest_register']
                    if student_register:
                        # Get the grade_level using the 'grade_level' field in Gradelevels
                        grade_level_obj = SMODEL.Gradelevels.objects.filter(ref=student_register.grade_level).first()
                        if grade_level_obj:
                            info['grade_level'] = grade_level_obj.grade_level  # Use the grade_level name from Gradelevels model
                        else:
                            info['grade_level'] = "N/A"  # Default if no matching grade level found
                    else:
                        info['grade_level'] = "N/A"  # Default if no register found

                if student_info:
                    print("Student info with latest registers and grade levels:", student_info)
                else:
                    print("No StudentRegister records found for these students.")
            else:
                print("No students found for this parent.")
        else:
            print("No student records associated with this parent.")
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")
        return render(request, 'school/error.html', {'message': 'An unexpected error occurred.'})

    # Return the data to the template
    return render(request, 'school/student_list.html', {
        'students': students,
        'student_info': student_info,
    })






# @login_required(login_url='studentlogin')
# @user_passes_test(is_student)
# def student_soa(request):
#     user_email = request.user.email
#     try:
        
#         student_ref = 131 # Change this as needed
        
#         parent = modelsv2.Parent.objects.get(email_address=user_email)
#         parentstudent = modelsv2.Parentstudent.objects.get(gid=parent.pid)
        
#         student_ref = parentstudent.stud_id
            
#     except Parent.DoesNotExist:
#         # Handle the case where the parent does not exist
#         print("Parent not found for the given email.")
#     except Parentstudent.DoesNotExist:
#         # Handle the case where the parent-student relationship does not exist
#         print("No student found for this parent.")
#     except Parentstudent.MultipleObjectsReturned:
#         # Handle the case where multiple students are found for the parent
#         print("Multiple students found for this parent.")
#     except Exception as e:
#         # Catch-all for other exceptions
#         print(f"An unexpected error occurred: {str(e)}")
    
      
#     try:
#         # Retrieve start and end dates from query parameters or hardcoded for now
#         start_date = '2024-06-01'
#         end_date = '2024-12-01'

#         # Retrieve student and register info
#         student = modelsv2.Student.objects.get(ref=student_ref)
#         studentsr = modelsv2.Studentregister.objects.filter(stud_id=student_ref).first()

        
#         student_info = {
#             'lrn_no': student.lrn_no,
#             'name': f"{student.last_name}, {student.first_name} {student.middle_name or ''}",
#             'grade': studentsr.grade_level,
#             'voucher': student.if_voucher,
#             'strand': studentsr.strand,
#         }

#         print("Student Info:", student_info)

#         # Retrieve charges within the date range
#         queryset = modelsv2.Scharges.objects.filter(stud_id=student_ref, date__range=[start_date, end_date])
#         total_charges = 0
#         # Print each record in the QuerySet
#         if queryset.exists():
#             print("Records found:")
#             for charge in queryset:
#                 print(f"Charge ID: {charge.table_pk}, Amount: {charge.amount}, Date: {charge.date}, Student ID: {charge.stud_id}")
#                 total_charges += charge.amount
#         else:
#             print("No charges found for this student between these dates.")
#         print("total charges", total_charges)
        
#                 # Retrieve charges within the date range
#         queryset2 = modelsv2.Spayment.objects.filter(stud_id=student_ref, date__range=[start_date, end_date])
#         total_pd = 0
#         # Print each record in the QuerySet
#         if queryset2.exists():
#             print("Records found:")
#             for pd in queryset2:
#                 print(f"Charge ID: {pd.table_pk}, Amount: {pd.amount}, Date: {pd.date}, Student ID: {pd.stud_id}")
#                 total_pd += pd.amount
#         else:
#             print("No charges found for this student between these dates.")
#         print("total charges", total_pd)
        
        
        
        
        
        
#         # Retrieve transactions after the start date
#         charges = modelsv2.Scharges.objects.filter(stud_id=student_ref, date__gte=start_date).annotate(
#             c_amount=F('amount'),
#             p_amount=Value(0),
#             transaction_type=Value('Charge')
#         )

#         payments = modelsv2.Spayment.objects.filter(stud_id=student_ref, date__gte=start_date).annotate(
#             c_amount=Value(0),
#             p_amount=F('amount'),
#             full_description=Concat('description', Value(' (R:'), F('pk'), Value(')'), output_field=CharField()),
#             transaction_type=Value('Payment')
#         )

#         # Combine and order transactions by date
#         transactions = sorted(
#             list(charges) + list(payments),
#             key=lambda t: t.date
#         )
#         print(transactions)

#         # Calculate running balance
#         balance = 0  # Initialize balance; set previous balance if applicable
#         for transaction in transactions:
#             balance += transaction.c_amount - transaction.p_amount
#             transaction.balance = balance

#         # Prepare context data
#         context = {
#             'transactions': transactions,
#             'student_info': student_info,
#             'previous_balance': balance,  # Set this according to your logic
#             'net_balance': balance,
#         }

#         # Render to the template
#         return render(request, 'school/student_soa.html', context)

#     # except student.DoesNotExist:
#     #     print(f"Student with ref {student_ref} not found")
#     #     return render(request, 'school/error.html', {'message': 'Student not found.'})  # Handle error gracefully
#     # except Studentregister.DoesNotExist:
#     #     print(f"Student register for student with ref {student_ref} not found")
#     #     return render(request, 'school/error.html', {'message': 'Student register not found.'})  # Handle error gracefully
#     except Exception as e:
#         print(f"An unexpected error occurred: {str(e)}")
#         return render(request, 'school/error.html', {'message': 'An unexpected error occurred.'})  # Handle unexpected errors
    
    
@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_grade(request):  # Accept student_ref as a parameter
    user_email = request.user.email
    student_ref = request.GET.get('student_ref')  # Get the student_ref from the URL parameters
    print(student_ref)
    try:
     
        # Get distinct grade levels the student is registered for
        # grade_levels = modelsv2.Studentregister.objects.filter(stud_id=student_ref).values('grade_level').distinct()
        grade_levels = SMODEL.Studentregister.objects.filter(stud_id=student_ref)
        grade_level_objects = []
        for grade_level in grade_levels:
            grade_level_obj = SMODEL.Gradelevels.objects.get(ref=grade_level.grade_level)
            grade_level_objects.append(grade_level_obj)
            print(grade_level_obj.grade_level)

        # Initialize data structures
        grades_data = []
        grading_periods = set()

        # Check if a grade level is selected via GET request
        selected_grade = request.GET.get('grade_level')
        print(selected_grade)
        # Determine which grades to fetch based on the selected grade level
        if selected_grade and selected_grade != "all":
            # Filter grades by the selected grade level
            grades = SMODEL.Grade.objects.filter(
                stud_id=student_ref,
                subject_code__grade_level__ref__grade_level=selected_grade  # Filter by grade level in the Subject model
            ).select_related('subject_code')
        else:
            # Get all grades for the student if 'all' is selected or no filter is applied
            grades = modelsv2.Grade.objects.filter(stud_id=student_ref).select_related('subject_code')

        # Prepare the grades data
        for grade in grades:
            subject = grade.subject_code  # Access the related Subject object directly
            subject_name = subject.sub_name if subject else "Unknown"  # Handle missing subjects
            grading_periods.add(grade.grading_period)  # Collect unique grading periods

            grades_data.append({
                'subject_code': grade.subject_code.subject_code,  # Include subject code
                'subject_name': subject_name,
                'stud_grade': grade.stud_grade,
                'gradelevel': grade.subject_code.grade_level,
                'grading_period': grade.grading_period,
            })

        # Pivot grades data for each subject and grading period
        pivoted_grades_data = {}
        for grade in grades_data:
            subject_name = grade['subject_name']
            grading_period = grade['grading_period']
            stud_grade = grade['stud_grade']

            if subject_name not in pivoted_grades_data:
                pivoted_grades_data[subject_name] = {'subject_name': subject_name}

            # Add grade for the grading period
            pivoted_grades_data[subject_name][grading_period] = stud_grade
        
        # Convert pivoted data to a list of dictionaries
        grades_data_list = list(pivoted_grades_data.values())
        grading_periods_list = sorted(list(grading_periods))  # Sort grading periods for consistency

        # # Check if it's an AJAX request (if it's an AJAX request, return JSON data)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'grades': grades_data_list,
                'grading_periods': grading_periods_list
            })

        # Prepare context for the initial page load
        context = {
            'grade_levels': grade_level_objects,
            'grades': grades_data_list,
            'grading_periods': grading_periods_list,
            'student_ref': student_ref,
        }

        return render(request, 'school/student_grade.html', context)

    except modelsv2.Parent.DoesNotExist:
        return render(request, 'school/error.html', {'message': 'Parent not found.'})
    except modelsv2.Parentstudent.DoesNotExist:
        return render(request, 'school/error.html', {'message': 'Student not associated with this parent.'})
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")
        return render(request, 'school/error.html', {'message': 'An unexpected error occurred.'})



# for aboutus and contact ussssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss
def aboutus_view(request):
    return render(request,'school/aboutus.html')

def contactus_view(request):
    sub = forms.ContactusForm()
    if request.method == 'POST':
        sub = forms.ContactusForm(request.POST)
        if sub.is_valid():
            email = sub.cleaned_data['Email']
            name=sub.cleaned_data['Name']
            message = sub.cleaned_data['Message']
            send_mail(str(name)+' || '+str(email),message,settings.EMAIL_HOST_USER, settings.EMAIL_RECEIVING_USER, fail_silently = False)
            return render(request, 'school/contactussuccess.html')
    return render(request, 'school/contactus.html', {'form':sub})



def save_sql(request):
    if request.method == "POST":
        
        # try:
        #     modelsv2.Scharges.objects.create(
        #         stud_id='STUD123',
        #         date='2024-11-18',
        #         description='Test',
        #         amount=100,
        #         atm_balance=500,
        #     )
        #     print("Record added successfully.")
        # except Exception as e:
        #     print(f"An error occurred: {e}")


        
        
        sql_statements = request.POST.get('sql_statements')

        # Split SQL statements into individual statements
        insert_statements = sql_statements.strip().split(";")
        modelsv2.Scharges.objects.all().delete()
        modelsv2.Spayment.objects.all().delete()
        for statement in insert_statements:
            # print(f"Processing statement: {statement.strip()}")
            if statement.strip():  # Skip empty statements
                try:
                    print(statement.strip().lower())
                    parse_and_save(statement.strip())
                except (ValueError, IntegrityError) as e:
                    print(f"Error: {e}")
                    return render(request, 'error.html', {'error_message': f'Error processing statement: {e}'})
                except Exception as e:
                    print(f"Unexpected error: {e}")
                    return render(request, 'error.html', {'error_message': f'Unexpected error: {e}'})

        return redirect('admin-dashboard')  # Redirect or show success page after saving

    return render(request, 'school/admin_savedb.html')  # If GET request, show the form


def parse_and_save(statement):
    print(f"Processing statement: {statement}")

    # Check for 'scharges' table insert statement
    if "insert into scharges" in statement.lower():
        # Clear all existing records in scharges table
        # modelsv2.Scharges.objects.all().delete()
        # print("All records in scharges table have been deleted.")

        # Use regular expression to extract values inside parentheses
        values_index = statement.lower().find("values")  # Find position of the 'VALUES' keyword

        # Slice the string to get everything after the 'VALUES' keyword and the following opening parenthesis
        values_text = statement[values_index + len("VALUES "):]  # Skip past the 'VALUES ' keyword
        values_text = values_text.strip()

        # Clean up any extra spaces or quotes, but retain the structure of the values
        # Split the values by comma, ensuring quoted strings remain intact
        values = []
        temp_value = ""
        inside_quotes = False

        for char in values_text:
            if char == "'" and not inside_quotes:
                inside_quotes = True
                temp_value += char
            elif char == "'" and inside_quotes:
                inside_quotes = False
                temp_value += char
                values.append(temp_value.strip())  # Add the value and reset
                temp_value = ""
            elif char == ',' and not inside_quotes:
                if temp_value.strip():
                    values.append(temp_value.strip())  # Add the value and reset
                temp_value = ""
            else:
                temp_value += char

        # Handle any leftover value (in case the last value is not followed by a comma)
        if temp_value.strip():
            values.append(temp_value.strip())

        print("Raw values extracted:", values)

        if values:
            if len(values) != 6:
                print(f"Error: Expected 6 values, but found {len(values)}")
                raise ValueError("Invalid number of values in INSERT statement")

            table_pk = values[0]
            stud_id = values[1].strip("'")
            date = values[2]
            description = values[3].strip("'")

            # Debugging: Print the raw amount and ATM Balance before conversion
            print(f"Amount before conversion: {values[4]}")
            print(f"ATM Balance before conversion: {values[5]}")

            # Convert the date to YYYY-MM-DD format
            try:
                date = date.strip("'")
                date = datetime.strptime(date, "%Y-%m-%d %H:%M:%S").date()  # Convert to date object
            except ValueError as e:
                print(f"Error converting date: {date}")
                raise e

            try:
                # Use regex to remove all non-numeric characters
                clean_amount = re.sub(r'\D', '', values[4])  # \D matches anything that is not a digit

                amount = int(clean_amount)  # Convert to integer
            except ValueError as e:
                print(f"Error converting amount: {values[4]} (must be a number)")
                raise e

            try:
                clean_amount = re.sub(r'\D', '', values[5])
                atm_balance = int(clean_amount)  # Convert to integer
            except ValueError as e:
                print(f"Error converting atm_balance: {values[5]} (must be a number)")
                raise e

            # Additional validation before creating the record
            if stud_id and date and description and amount >= 0 and atm_balance >= 0:
                # Only create the record if all conditions are met
                modelsv2.Scharges.objects.create(
                    stud_id=stud_id,
                    date=date,
                    description=description,
                    amount=amount,
                    atm_balance=atm_balance,
                )
                print("Record successfully added to scharges.")
            else:
                print(f"Validation failed for stud_id: {stud_id}, date: {date}, description: {description}, amount: {amount}, atm_balance: {atm_balance}")

    # Check for 'spayment' table insert statement
    elif "insert into spayment" in statement.lower():
        # Clear all existing records in spayment table
        # modelsv2.Spayment.objects.all().delete()
        # print("All records in spayment table have been deleted.")

        # Use regular expression to extract values inside parentheses
        values_index = statement.lower().find("values")  # Find position of the 'VALUES' keyword

        # Slice the string to get everything after the 'VALUES' keyword and the following opening parenthesis
        values_text = statement[values_index + len("VALUES "):]  # Skip past the 'VALUES ' keyword
        values_text = values_text.strip()

        # Clean up any extra spaces or quotes, but retain the structure of the values
        # Split the values by comma, ensuring quoted strings remain intact
        values = []
        temp_value = ""
        inside_quotes = False

        for char in values_text:
            if char == "'" and not inside_quotes:
                inside_quotes = True
                temp_value += char
            elif char == "'" and inside_quotes:
                inside_quotes = False
                temp_value += char
                values.append(temp_value.strip())  # Add the value and reset
                temp_value = ""
            elif char == ',' and not inside_quotes:
                if temp_value.strip():
                    values.append(temp_value.strip())  # Add the value and reset
                temp_value = ""
            else:
                temp_value += char

        # Handle any leftover value (in case the last value is not followed by a comma)
        if temp_value.strip():
            values.append(temp_value.strip())

        print("Raw values extracted:", values)

        if values:
            if len(values) != 7:
                print(f"Error: Expected 7 values, but found {len(values)}")
                raise ValueError("Invalid number of values in INSERT statement")

            table_pk = values[0]
            stud_id = values[2].strip("'")
            date = values[3]
            description = values[4].strip("'")

            # Debugging: Print the raw amount and ATM Balance before conversion
            print(f"Amount before conversion: {values[4]}")
            print(f"ATM Balance before conversion: {values[5]}")

            # Convert the date to YYYY-MM-DD format
            try:
                date = date.strip("'")
                date = datetime.strptime(date, "%Y-%m-%d %H:%M:%S").date()  # Convert to date object
            except ValueError as e:
                print(f"Error converting date: {date}")
                raise e

            try:
                clean_amount = re.sub(r'\D', '', values[5])
                amount = int(clean_amount)
            except ValueError as e:
                print(f"Error converting amount: {values[5]} (must be a number)")
                raise e

            try:
                clean_amount = re.sub(r'\D', '', values[6])
                atm_balance = int(clean_amount)
            except ValueError as e:
                print(f"Error converting atm_balance: {values[6]} (must be a number)")
                raise e

            # Additional validation before creating the record
            if stud_id and date and description and amount >= 0 and atm_balance >= 0:
                # Only create the record if all conditions are met
                modelsv2.Spayment.objects.create(
                    table_pk=table_pk,
                    stud_id=stud_id,
                    date=date,
                    description=description,
                    amount=amount,
                    atm_balance=atm_balance,
                )
                print("Record successfully added to spayment.")
            else:
                print(f"Validation failed for table_pk: {table_pk}, stud_id: {stud_id}, date: {date}, description: {description}, amount: {amount}, atm_balance: {atm_balance}")
    else:
        print("Statement does not match any known table pattern.")
        
        
        
        
        

# @login_required(login_url='studentlogin')
# @user_passes_test(is_student)
# def student_dashboard_view(request):
#     dict={
    
#     'total_course':QMODEL.Course.objects.all().count(),
#     'total_question':QMODEL.Question.objects.all().count(),
#     }
#     return render(request,'/student_dashboard.html',context=dict)

# @login_required(login_url='studentlogin')
# @user_passes_test(is_student)
# def student_exam_view(request):
#     courses=QMODEL.Course.objects.all()
#     return render(request,'student/student_exam.html',{'courses':courses})


@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_exam_view(request):
    
    student_ref = '131'  # Get the student_ref from the URL parameters
    print(student_ref)    
    # Get distinct grade levels the student is registered for
    grade_levels = modelsv2.Studentregister.objects.filter(stud_id=student_ref).values('grade_level').distinct()
    # Filter courses based on the student's grade level
        # If multiple grade levels exist, you can choose how to handle it (e.g., the first grade level or filter by all)
    if grade_levels:
        grade_level = grade_levels[0]['grade_level']  # Choose the first grade level (if there are multiple)
        print(grade_level)
        # Filter courses based on the student's grade level
        courses = QMODEL.Course.objects.filter(grade_level=grade_level)
    else:
        # No grade level found for student
        courses = []

    # Pass the filtered courses to the template
    return render(request, 'student/student_exam.html', {'courses': courses})



@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def take_exam_view(request,pk):
    student_ref = '131'  # Get the student_ref from the URL parameters
    print(student_ref)    
    
    # Get distinct grade levels the student is registered for
    grade_levels = modelsv2.Studentregister.objects.filter(stud_id=student_ref).values('grade_level').distinct()
    grade_level = grade_levels[0]['grade_level']
    # Filter courses based on the student's grade level
    course=QMODEL.Course.objects.get(id=pk,grade_level=grade_level)
    total_questions=QMODEL.Question.objects.all().filter(course=course).count()
    questions=QMODEL.Question.objects.all().filter(course=course)
    total_marks=0
    for q in questions:
        total_marks=total_marks + q.marks
    
    return render(request,'student/take_exam.html',{'course':course,'total_questions':total_questions,'total_marks':total_marks})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def start_exam_view(request,pk):
    course=QMODEL.Course.objects.get(id=pk)
    questions=QMODEL.Question.objects.all().filter(course=course)
    if request.method=='POST':
        pass
    response= render(request,'student/start_exam.html',{'course':course,'questions':questions})
    response.set_cookie('course_id',course.id)
    return response


@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def calculate_marks_view(request):
    if request.COOKIES.get('course_id') is not None:
        course_id = request.COOKIES.get('course_id')
        course=QMODEL.Course.objects.get(id=course_id)
        
        total_marks=0
        questions=QMODEL.Question.objects.all().filter(course=course)
        for i in range(len(questions)):
            
            selected_ans = request.COOKIES.get(str(i+1))
            actual_answer = questions[i].answer
            if selected_ans == actual_answer:
                total_marks = total_marks + questions[i].marks
        student = SMODEL.Student.objects.get(ref=request.user.id)
        result = QMODEL.Result()
        result.marks=total_marks
        result.exam=course
        result.student=student
        result.save()

        return HttpResponseRedirect('view-result')



@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def view_result_view(request):
    courses=QMODEL.Course.objects.all()
    return render(request,'student/view_result.html',{'courses':courses})
    

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def check_marks_view(request,pk):
    course=QMODEL.Course.objects.get(id=pk)
    student = models.Student.objects.get(user_id=request.user.id)
    results= QMODEL.Result.objects.all().filter(exam=course).filter(student=student)
    return render(request,'student/check_marks.html',{'results':results})

@login_required(login_url='studentlogin')
@user_passes_test(is_student)
def student_marks_view(request):
    courses=QMODEL.Course.objects.all()
    return render(request,'student/student_marks.html',{'courses':courses})
  
